import tkinter
from tkinter import messagebox
import os
import openpyxl

# Defines a function named 'enter_data' that allows the button to enter data into the defined spreadsheet
def enter_data():

        # Recipe info
        Recipe_Name = recipe_name_entry.get()
        Ingredients = ingredient_entry.get()
        Measurements = measure_entry.get()
        Method = method_entry.get()
        Servings = servings_spinbox.get()

        # if statements fulfilling requirement that data will not be entered into spreadsheet if fields are left empty
        # if fields are indeed left empty the appropriate error message box is presented to the user
        if recipe_name_entry.index("end") == 0:
            tkinter.messagebox.showwarning(title='Error', message='Recipe must have a name')
        if ingredient_entry.index("end") == 0:
            tkinter.messagebox.showwarning(title='Error', message='Recipe must have ingredients')
        if measure_entry.index("end") == 0:
            tkinter.messagebox.showwarning(title='Error', message='Recipe must have measurements')
        if method_entry.index("end") == 0:
            tkinter.messagebox.showwarning(title='Error', message='Recipe must have a method')

        if Recipe_Name and Measurements:
            serves = servings_spinbox.get()

            # recipe info

            print("Recipe Name:", Recipe_Name)
            print("Ingredients: ", Ingredients)
            print("Measurements: ", Measurements)
            print("Method: ", Method)
            print("Servings: ", Servings)

            #where will the data file be saved
            filepath = "C:\\Users\\Kille\\PycharmProjects\\pythonProject\\SAC FORM\\SAC DATABASE.xlsx"

            # if the data file cant be found it will simply create a new file (only used for first time setup)
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                heading = ["Recipe Name", "Ingredients", "Measurements", "Method", "Servings"]
                sheet.append(heading)
                workbook.save(filepath)
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            sheet.append([Recipe_Name, Ingredients, Measurements, Method, Servings])
            workbook.save(filepath)

#defines a function named clear form
def clear_form():
    #clearing user info
    recipe_name_entry.delete(0, tkinter.END)
    ingredient_entry.delete(0, tkinter.END)
    measure_entry.delete(0, tkinter.END)
    method_entry.delete(0, tkinter.END)
    servings_spinbox.delete(0, tkinter.END)

### GETTING EXCEL DATA AND PRINTING IN CONSOLE

def get_form():
    heading = ["Recipe Name", "Ingredients", "Measurements", "Method", "Servings"]
    if not os.path.exists("C:\\Users\\Kille\\PycharmProjects\\pythonProject\\SAC FORM\\SAC DATABASE.xlsx"):
        tkinter.messagebox.showwarning(title='Error', message='No data inside spreadsheet or spreadsheet could not be found')
        return
    workbook = openpyxl.load_workbook("C:\\Users\\Kille\\PycharmProjects\\pythonProject\\SAC FORM\\SAC DATABASE.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        count = 0
        print('--------------') # This has been added so that in the console you can differentiate recipes that have just been added to recipes that are printed out.
        # these for in and if in statements make sure the code does not print the heading row of the database
        for cell in row:
            if cell in heading:
                continue
            print(f'{heading[count]}: {cell}')
            count += 1

window = tkinter.Tk() # Creates the main window object.
window.title("Recipe Entry Form") # Sets the title of the window

window.iconbitmap("down.ico") # Sets the icon of the window

frame = tkinter.Frame(window) # creates a frame widget within the main window.
frame.pack() # packs the frame widget to make it visible within the window.

# Creates and configues various GUI elements
# Lables entry fields comboboxes spinboxes check buttons
# and arranges them within the frames using the 'grid()' method

# Saving user info
user_info_frame = tkinter.LabelFrame(frame, text="Recipe information")
user_info_frame.grid(row=0, column=0, padx=20, pady=10)

recipe_name_label = tkinter.Label(user_info_frame, text="Recipe Name")
recipe_name_label.grid(row=0, column=0)
ingredientlabel = tkinter.Label(user_info_frame, text="Ingredients")
ingredientlabel.grid(row=0, column=1)
measurements_label = tkinter.Label(user_info_frame, text="Measurements")
measurements_label.grid(row=0, column=2)

method_label = tkinter.Label(user_info_frame, text="Method")
method_label.grid(row=2, column=0, padx=25)

recipe_name_entry = tkinter.Entry(user_info_frame)
ingredient_entry = tkinter.Entry(user_info_frame)
measure_entry = tkinter.Entry(user_info_frame)
method_entry = tkinter.Entry(user_info_frame)
recipe_name_entry.grid(row=1, column=0)
ingredient_entry.grid(row=1, column=1)
measure_entry.grid(row =1, column=2)
method_entry.grid(row =3, column= 0)

servings_label = tkinter.Label(user_info_frame, text="Servings")
servings_spinbox = tkinter.Spinbox(user_info_frame, from_=1, to=10000) # Sets the min and max values of the spinbox to 1-10000
servings_label.grid(row=2, column=1)
servings_spinbox.grid(row=3, column=1)

#configures padding for the widgets within the 'user_info_frame'.z
for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

#creates a button widget with the text "Save data to MS excel file" and associates it with the 'enter_data' function
button1 = tkinter.Button(frame, text="Save data to MS excel file", command=enter_data, bg="#00E4FF")
#button - arranges the button widget within the frame using the 'grid()' method.
button1.grid(row=3, column=0, sticky="news", padx=20, pady=10)

#creates a button widget with the text "Clear Form Data" and associates it with the 'clear_form' function
button2 = tkinter.Button(frame, text="Clear Form Data", command=clear_form, bg="#FF4200")
#button - arranges the button widget within the frame using the 'grid()' method.
button2.grid(row=5, column=0, sticky="news", padx=20, pady=10)

# creates a button widget with the text "Print Spreadsheet Data" and associates it with the 'get_form' function
button3 = tkinter.Button(frame, text="Print Spreadsheet Data", command=get_form, bg="#00FF83")
#button - arranges the button widget within the frame using the 'grid()' method.
button3.grid(row=6, column=0, sticky="news", padx=20,pady=10)

#starts the main event loop, which listens for events and keeps the gui window open until it is closed
window.mainloop()